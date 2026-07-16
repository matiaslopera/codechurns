import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.actions import ACTION_DISCOUNT, ACTION_OUTREACH, ACTION_REMINDER
from core.i18n import t

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Row color communicates risk severity at a glance (red = most urgent).
_ROW_FILLS = {
    ACTION_OUTREACH: PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    ACTION_DISCOUNT: PatternFill(start_color="FDE8CC", end_color="FDE8CC", fill_type="solid"),
    ACTION_REMINDER: PatternFill(start_color="FFF6CC", end_color="FFF6CC", fill_type="solid"),
    None: PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
}

_COLUMN_WIDTHS = [20, 16, 18, 12, 50, 32, 14, 14, 40]


def build_workbook(export_rows, lang="es"):
    """export_rows: list of dicts with keys client, normal_cycle_days,
    days_since_last_visit, risk_ratio, explanation, action_key, action_label,
    decision, feedback, final_action_text. Returns a BytesIO ready for a
    download button."""
    wb = Workbook()
    ws = wb.active
    ws.title = t("dashboard_title", lang)[:31]

    headers = [
        t("export_col_client", lang),
        t("export_col_cycle", lang),
        t("export_col_days_since", lang),
        t("export_col_ratio", lang),
        t("export_col_explanation", lang),
        t("export_col_suggested_action", lang),
        t("export_col_decision", lang),
        t("export_col_feedback", lang),
        t("export_col_final_action", lang),
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in export_rows:
        ws.append(
            [
                row["client"],
                row["normal_cycle_days"],
                row["days_since_last_visit"],
                row["risk_ratio"],
                row["explanation"],
                row["action_label"],
                t(f"decision_{row['decision']}", lang),
                t(f"feedback_{row['feedback']}", lang),
                row["final_action_text"],
            ]
        )
        fill = _ROW_FILLS.get(row.get("action_key"), _ROW_FILLS[None])
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
