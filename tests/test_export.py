from openpyxl import load_workbook

from core.actions import ACTION_OUTREACH
from core.export import build_workbook


def _sample_rows():
    return [
        {
            "client": "Carlos Ruiz",
            "normal_cycle_days": 14.0,
            "days_since_last_visit": 66,
            "risk_ratio": 4.71,
            "explanation": "Suele venir cada 14 días...",
            "action_key": ACTION_OUTREACH,
            "action_label": "Llamada personal + oferta",
            "decision": "approved",
            "feedback": "useful",
            "final_action_text": "Llamar y ofrecer 20% de descuento.",
        }
    ]


def test_build_workbook_has_header_and_data_row():
    buffer = build_workbook(_sample_rows(), lang="es")
    wb = load_workbook(buffer)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Cliente"
    assert ws.cell(row=2, column=1).value == "Carlos Ruiz"
    assert ws.cell(row=2, column=7).value == "Aprobada"
    assert ws.cell(row=1, column=8).value == "Feedback"
    assert ws.cell(row=2, column=8).value == "Útil"


def test_build_workbook_applies_row_fill_by_action():
    buffer = build_workbook(_sample_rows(), lang="es")
    wb = load_workbook(buffer)
    ws = wb.active
    fill = ws.cell(row=2, column=1).fill
    assert fill.start_color.rgb == "00F8D7DA"


def test_build_workbook_empty_rows_only_has_header():
    buffer = build_workbook([], lang="es")
    wb = load_workbook(buffer)
    ws = wb.active
    assert ws.max_row == 1


def test_build_workbook_localized_headers_english():
    buffer = build_workbook(_sample_rows(), lang="en")
    wb = load_workbook(buffer)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Customer"
    assert ws.cell(row=2, column=7).value == "Approved"
    assert ws.cell(row=2, column=8).value == "Useful"


def test_build_workbook_unmarked_and_not_useful_feedback():
    rows = _sample_rows()
    rows[0]["feedback"] = "not_useful"
    buffer = build_workbook(rows, lang="es")
    ws = load_workbook(buffer).active
    assert ws.cell(row=2, column=8).value == "No útil"

    rows[0]["feedback"] = "unmarked"
    buffer = build_workbook(rows, lang="es")
    ws = load_workbook(buffer).active
    assert ws.cell(row=2, column=8).value == "Sin marcar"
